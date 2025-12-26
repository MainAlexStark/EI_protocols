from email import header
import json
import os
from pathlib import Path
import random
import re
import sys
import time

from PyQt6.QtCore import QSize, Qt
from PyQt6.QtCore import QAbstractTableModel, Qt, QVariant
from PyQt6.QtCore import QObject, pyqtSignal, QThread
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QPushButton,
    QVBoxLayout, QLineEdit, QLabel, QFileDialog, QWidget, QCheckBox, QDialog, QHBoxLayout ,QMessageBox, QPlainTextEdit, QProgressBar
)
from PyQt6.QtWidgets import QTableView
from openpyxl import load_workbook
import pandas as pd

from EI_protocols_utils.utils.constants import *
from EI_protocols_utils.utils.weather import get_weather, add_weather
from EI_protocols_utils.utils.exchanges import RequiredFieldsError, RowError
from EI_protocols_utils.utils.models import Journal, WaterMeterProtocol
from EI_protocols_utils.utils.settings import settings
from EI_protocols_utils.utils.user_info import save_paths, load_paths
data = load_paths(filename=settings.user_info_path)

# Работа со средним выполнением протоколов
TIME_STATS_FILE = Path(settings.protocol_times_path)
MAX_HISTORY = 50  # будем хранить 50 последних значений

if TIME_STATS_FILE.exists():
    with open(TIME_STATS_FILE, "r") as f:
        times_data=json.load(f)
        times = times_data['times']
else:
    times = []

class ProtocolWorker(QObject):
    progress = pyqtSignal(int)         # процент
    message = pyqtSignal(str)          # лог в консоль
    finished = pyqtSignal(list, list, list)  # completed, errors, notcomplited
    eta = pyqtSignal(float)             # сигнал оставшегося времени в секундах
    
    required_fields = [0, 2, 5, 7, 9, 10, 12, 13, 21, 32, 35, 44, 45]

    def __init__(self, workbook, from_row, to_row, protocols_path, journal_path):
        super().__init__()
        self.workbook = workbook
        self.wsheet = workbook[JOURNAL_WORKSHEET]
        self.from_row = from_row
        self.to_row = to_row
        self.protocols_path = protocols_path
        self.journal_path = journal_path
        
    def validate_row(self, row: list) -> list:
        # Выводим пропущенные обязательные поля
        missing_fields = []
        for index in self.required_fields:
            if not row[index]:
                missing_fields.append(index)
        
        cells_names = {}
        first_row = self.wsheet[1] # Нумерация в openpyxl с 1
        for index in self.required_fields:
            field_name = first_row[index].value
            cells_names[index] = field_name
        
        if missing_fields:
            # Получаем названия пропущенных полей
            missing_field_names = [cells_names[index] for index in missing_fields]
            # Формируем человекочитаемую строку
            missing_field_list = ", ".join(missing_field_names)
            raise RequiredFieldsError(f"Пропущены поля: {missing_field_list}", missing_fields)

        row[1] = '1' if not row[1] else str(row[1])
        
        # Если не указана погода, то создаем или берем из журнала погоды
        if not row[14] or not row[15] or not row[16]:
            date_str = row[9].strftime("%d.%m.%Y") if type(row[9]) is not str else row[9]
            weather = get_weather(date_str)
            if weather and (not row[14] or not row[15] or not row[16]):
                row[14] = weather["temperature"] + ' °C'
                row[15] = weather["pressure"] + ' кП'
                row[16] = weather["humidity"] + ' %'
                if not get_weather(date_str):
                    add_weather(date_str, row[14], row[15], row[16])  # Обновляем журнал погоды
            else:
                # Если погоды нет, то генерируем случайные значения и сохраняем в журнал погоды
                temperature = str(round(random.uniform(settings.temperatures["min"], settings.temperatures["max"]), 1))
                pressure = str(round(random.uniform(settings.pressure["min"], settings.pressure["max"]), 1))
                humidity = str(round(random.uniform(settings.humdity["min"], settings.humdity["max"]), 1))
                add_weather(date_str, temperature, pressure, humidity)
                row[14] = temperature + ' °C'
                row[15] = pressure + ' кП'
                row[16] = humidity + ' %'

        row[34] = '1' if not row[34] else str(row[34])
        
        if not row[47]: row[47] = "Частное лицо"

        return row

    def run(self):
        global times
        errors = []
        completed = []
        not_completed = []

        total_count = self.to_row - self.from_row 

        for i, row in enumerate(self.wsheet.iter_rows(min_row=self.from_row, max_row=self.to_row, values_only=False)):
            try:
                # Преобразуем к значениям для работы
                values = [cell.value for cell in row]

                # for index, value in enumerate(values):
                #     print(f"{index}: {value}")

                # Проверяем и дополняем данные
                self.validate_row(values)
                
                temperature_str = re.sub(r"[^0-9.]", "", str(values[14]).replace(',', '.'))
                pressure_str = re.sub(r"[^0-9.]", "", str(values[15]).replace(',', '.'))
                humidity_str = re.sub(r"[^0-9.]", "", str(values[16]).replace(',', '.'))
                readings_str = re.sub(r"[^0-9.]", "", str(values[45]).replace(',', '.'))
                
                temperature_float = round(float(temperature_str), 1)
                pressure_float = round(float(pressure_str), 1)
                humidity_float = round(float(humidity_str), 1)
                readings_float = round(float(readings_str), 3)
                
                second_number = None
                if values[41]:
                    match = re.search(r'\(([^-]+)-([^)]+)\)', values[41]) 
                    second_number = float(match.group(2).replace(',', '.'))

                # Создаем протокол
                protocol = WaterMeterProtocol(
                    dir_path=self.protocols_path,
                    tab_number=settings.tab_numbers.get(values[35], values[0].split('-')[2]),
                    protocol_number=values[0].split('-')[-1],
                    date=values[9].strftime("%d.%m.%Y") if type(values[9]) is not str else values[9],
                    next_date=values[10].strftime("%d.%m.%Y") if type(values[10]) is not str else values[10],
                    SI_numbers=values[12],
                    suitability=False if values[13]=="Непригодно" else True,
                    reasons_for_unsuitability=values[38],
                    name=values[5],
                    number=values[7],
                    register_number=values[2],
                    year=int(values[44]),
                    owner=values[47] if values[47] else "Частное лицо",
                    address=values[32],
                    temperature=temperature_float,
                    pressure=pressure_float,
                    humidity=humidity_float,
                    readings=readings_float,
                    unit_type=values[48],
                    range=str(second_number)
                )
                
                start = time.perf_counter()
                xlsx_path, pdf_path = protocol.create()
                end = time.perf_counter()

                elapsed = end - start
                times.append(elapsed)
                if len(times) > MAX_HISTORY:
                    times = times[-MAX_HISTORY:]

                # сохраняем обратно
                with open(TIME_STATS_FILE, "w") as f:
                    json.dump({"times": times}, f)
                    
                avg_time = sum(times) / len(times)
                remaining = avg_time * (total_count - i)

                self.eta.emit(remaining)               # отправляем оставшееся время в диалог
                self.progress.emit(int(i / total_count * 100))

                completed.append((xlsx_path, pdf_path))
                self.message.emit(f"✅Создан протокол: {xlsx_path}")

                ready_protocol = load_workbook(filename=xlsx_path, data_only=True) # Без data_only=False будут формулы
                ready_wsheet = ready_protocol[WATER_METER_PROTOTOCOL_SEETNAME]
                
                if values[41]:
                    pass
                else:
                    if "Измерения на расходе Qнаиб , л/ч" in str(ready_wsheet["B55"].value):
                        consumption=max(float(ready_wsheet["AC55"].value), float(ready_wsheet["AC54"].value), float(ready_wsheet["AC53"].value))
                    elif "Измерения на расходе Qнаиб , л/ч" in ready_wsheet["B59"].value:
                        consumption=max(float(ready_wsheet["AC61"].value), float(ready_wsheet["AC60"].value), float(ready_wsheet["AC59"].value))
                    else:
                        raise Exception("Не удалось определить максимальный расход из протокола"\
                            "Убедитесь, что шаблон протокола содержит результаты расхода измерений в B53-B55, или AC59-AC61")
                    
                    values[41] = f"Поверен в диапазоне расхода (0,03-{round(consumption, 3)}) м3/ч"
                    
                for i, cell in enumerate(row):
                    cell.value = values[i]
                    
            except Exception as e:
                errors.append(RowError(e, i+self.from_row))
                not_completed.append(i+self.from_row)
                self.message.emit(f"❌Ошибка: {e}, строка {i+self.from_row}\n")
                raise e
            except RequiredFieldsError as rfe:
                not_completed.append(i+self.from_row)
                errors.append(RowError(rfe, i+self.from_row))
                self.message.emit(f"❌Ошибка: {rfe}, строка {i+self.from_row}\n")
                
            self.workbook.save(filename=self.journal_path)

            # обновляем прогресс
            progress_percent = int(i / total_count * 100)
            self.progress.emit(progress_percent)

        self.finished.emit(completed, errors, not_completed)

class CreateProtocolDialog(QDialog):
    def __init__(self, journal_path, protocols_path, from_row, to_row):
        super().__init__()
        
        self.journal_path = journal_path
        self.protocols_path = protocols_path
        self.from_row = from_row
        self.to_row = to_row
        
        self.workbook = load_workbook(filename=self.journal_path)
        self.wsheet = self.workbook[JOURNAL_WORKSHEET]
        
        self.setWindowTitle("Создание протоколов")
        self.setFixedSize(QSize(500, 500))
        self.main_layout = QVBoxLayout(self)
        
        # ---- Консоль ----
        self.console = QPlainTextEdit()
        self.console.setPlainText("Начало создания протоколов...\n")
        self.console.setLineWrapMode(QPlainTextEdit.LineWrapMode.NoWrap)
        self.console.setFixedSize(QSize(480, 400))
        self.console.setReadOnly(True)
        self.main_layout.addWidget(self.console)
        
        # ---- Прогрессбар ----
        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        self.progress.setFixedSize(QSize(480, 30))
        self.main_layout.addWidget(self.progress)
        
        self.workbook = load_workbook(journal_path)
        self.worker_thread = QThread()
        self.worker = ProtocolWorker(self.workbook, from_row, to_row, protocols_path, journal_path)
        self.worker.moveToThread(self.worker_thread)
        
        # label
        self.eta_label = QLabel("Осталось ~0 сек.")
        self.main_layout.addWidget(self.eta_label)


        # сигналы
        self.worker_thread.started.connect(self.worker.run)
        self.worker.progress.connect(self.progress.setValue)
        self.worker.message.connect(self.console.appendPlainText)
        self.worker.finished.connect(self.on_finished)
        self.worker.finished.connect(self.worker_thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.worker_thread.finished.connect(self.worker_thread.deleteLater)
        self.worker.eta.connect(self.update_eta_label)

        self.worker_thread.start()
        
    def update_eta_label(self, remaining_seconds):
        mins, secs = divmod(int(remaining_seconds), 60)
        self.eta_label.setText(f"Осталось ~{mins} мин {secs} сек")
        
    def on_finished(self, completed, errors, not_completed):
        self.completed_protocols = completed
        self.error_protocols = errors
        self.console.appendPlainText("🧩Создание протоколов завершено!")        
        self.console.appendPlainText(f"🧩Выполнены успешно ({len(completed)}): \n")
        for item in completed:
            self.console.appendPlainText(f" - {item[0].name}, {item[1].name}")
        self.console.appendPlainText(f"\n⚠️С ошибками ({len(errors)}): ")
        for error in errors:
            self.console.appendPlainText(f" - {error}, строка {error.row_number}")
            
        if errors:
            self.question_label = QLabel(f"Возникли ошибки, сохранить протоколы?")
            self.yes_button = QPushButton("Да")
            self.yes_button.clicked.connect(self.yes_button_clicked)
            self.no_button = QPushButton("Нет")
            self.no_button.clicked.connect(self.no_button_clicked)

            self.main_layout.addWidget(self.question_label)
            self.main_layout.addWidget(self.yes_button)
            self.main_layout.addWidget(self.no_button)
        
    def yes_button_clicked(self):
        self.accept()
    def no_button_clicked(self):
        for item in self.completed_protocols:
            self.console.appendPlainText(f"Удаляем протоколы: {item[0].name}, {item[1].name}")
            os.remove(item[0])
            os.remove(item[1])
        # for item in self.error_protocols:
        #     os.remove(item[0])
        #     os.remove(item[1])
        self.accept()
                    

class SettingsDialog(QDialog):
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Настройки")
        self.setFixedSize(QSize(300, 200))
        layout = QVBoxLayout(self)
        
        self.app_settings = load_paths(filename=settings.app_settings_path)
        for key, value in self.app_settings.items():
            if isinstance(value, bool):
                checkbox = QCheckBox(f"{key}")
                checkbox.setChecked(value)
                checkbox.stateChanged.connect(self.update_setting(key))
                layout.addWidget(checkbox)
            
    def update_setting(self, key):
        def handler(state):
            self.app_settings[key] = bool(state)
            save_paths(self.app_settings, filename=settings.app_settings_path)
        return handler
    
    
    
class PandasModel(QAbstractTableModel):
    """Модель, адаптирующая pandas.DataFrame под Qt TableView."""
    
    def __init__(self, df):
        super().__init__()
        self._df = df

    def rowCount(self, parent=None):
        return len(self._df.index)

    def columnCount(self, parent=None):
        return len(self._df.columns)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return QVariant()

        if role == Qt.ItemDataRole.DisplayRole:
            value = self._df.iat[index.row(), index.column()]
            return str(value)

        return QVariant()

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role != Qt.ItemDataRole.DisplayRole:
            return QVariant()

        if orientation == Qt.Orientation.Horizontal:
            return str(self._df.columns[section])
        else:
            return str(self._df.index[section])

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("ЕИ Протоколы")

        # Создаём центральный виджетcd 
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)

        # Создаём layout для центрального виджета
        main_layout = QHBoxLayout()
        central_widget.setLayout(main_layout)

        # Секция создания протокола
        create_protocol_layout = QVBoxLayout()
        create_protocol_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self.journal_path_button = QPushButton("Путь к журналу")
        self.journal_path_button.setFixedSize(200, 40)
        create_protocol_layout.addWidget(self.journal_path_button)
        self.journal_path_button.clicked.connect(self.select_journal_path)
        
        self.journal_path_label = QLabel(data.get("journal_path", ""))
        create_protocol_layout.addWidget(self.journal_path_label)
        
        self.protocols_path_button = QPushButton("Путь к протоколам")
        self.protocols_path_button.setFixedSize(200, 40)
        create_protocol_layout.addWidget(self.protocols_path_button)
        self.protocols_path_button.clicked.connect(self.select_protocols_path)

        self.protocols_path_label = QLabel(data.get("protocols_path", ""))
        create_protocol_layout.addWidget(self.protocols_path_label)
        
        self.from_row_input = QLineEdit()
        self.from_row_input.setPlaceholderText("С какой строки")
        self.from_row_input.setFixedSize(200, 40)
        create_protocol_layout.addWidget(self.from_row_input)
        
        self.to_row_input = QLineEdit()
        self.to_row_input.setPlaceholderText("По какую строку")
        self.to_row_input.setFixedSize(200, 40)
        create_protocol_layout.addWidget(self.to_row_input)

        self.create_protocol_button = QPushButton("Создать протоколы")
        self.create_protocol_button.setFixedSize(200, 40)
        self.create_protocol_button.clicked.connect(self.create_protocols)
        create_protocol_layout.addWidget(self.create_protocol_button)

        self.settings_button = QPushButton("Настройки")
        self.settings_button.setFixedSize(200, 40)
        self.settings_button.clicked.connect(self.open_settings)
        create_protocol_layout.addWidget(self.settings_button)

        main_layout.addLayout(create_protocol_layout)
        
        
        ### Layout с таблицей
        self.table_layout = QVBoxLayout()
        
        self.table = QTableView()
        if data.get("journal_path", ""): self.load_excel_to_table(data["journal_path"])
        self.table_layout.addWidget(self.table)
        
        main_layout.addLayout(self.table_layout)


    ##################### Методы работы с таблицей
    def load_excel_to_table(self, path):
        try:
            df = pd.read_excel(path)

            # создаём модель
            self.model = PandasModel(df)
            self.table.setModel(self.model)
            self.table.selectionModel().selectionChanged.connect(self.on_selection_changed)
            
            self.table.scrollToBottom()

            # подгоняем колонки
            self.table.horizontalHeader().setStretchLastSection(True)
            self.table.resizeColumnsToContents()

            # Восстанавливаем сохранённые ширины
            settings_data = load_paths(filename=settings.app_settings_path)
            widths = settings_data.get("column_widths", {})

            for col_index in range(len(df.columns)):
                col_name = df.columns[col_index]
                if col_name in widths:
                    self.table.setColumnWidth(col_index, widths[col_name])

            # Слушаем изменения ширины пользователя
            header = self.table.horizontalHeader()
            header.sectionResized.connect(self.save_column_width)

        except Exception as e:
            print("Ошибка загрузки Excel:", e)
            
    def save_column_width(self, index, old_size, new_size):
        # Создаём словарь если нет
        settings_data = load_paths(filename=settings.app_settings_path)
        if "column_widths" not in settings_data:
            settings_data["column_widths"] = {}

        # Берём имя столбца по индексу
        column_name = self.model._df.columns[index]
        settings_data["column_widths"][column_name] = new_size

        save_paths(settings_data, filename=settings.app_settings_path)
        
    def get_selected_rows(self):
        indexes = self.table.selectionModel().selectedIndexes()
        rows = sorted(set(index.row() for index in indexes))
        return rows
    
    def on_selection_changed(self, selected, deselected):
        rows = self.get_selected_rows()
        self.from_row_input.setText(str(min(rows)+2) if rows else "")
        self.to_row_input.setText(str(max(rows)+2) if rows else "")
    ################


    #### Создание протоколов
        
    def create_protocols(self):
        journal_path = self.journal_path_label.text()
        protocols_path = self.protocols_path_label.text()
        from_row = self.from_row_input.text()
        to_row = self.to_row_input.text()
        
        if not journal_path or not protocols_path or not from_row or not to_row:
            QMessageBox.warning(self, "Осторожно", "Заполните все поля!")
            return
        
        dialog = CreateProtocolDialog(
            journal_path=journal_path,
            protocols_path=protocols_path,
            from_row=int(from_row),
            to_row=int(to_row)
        )
        dialog.exec()
        
        
    #############

    def select_journal_path(self):
        file, _ = QFileDialog.getOpenFileName(self, caption="Выберите файл", filter="Excel файлы (*.xlsx)")
        if file:
            self.journal_path_label.setText(file)
            data["journal_path"] = file
            save_paths(data, filename=settings.user_info_path)
            self.load_excel_to_table(file)
            
    def select_protocols_path(self):
        directory = QFileDialog.getExistingDirectory(self, caption="Выберите папку для протоколов")
        if directory:
            self.protocols_path_label.setText(directory)
            data["protocols_path"] = directory
            save_paths(data, filename=settings.user_info_path)
            
    def open_settings(self):
        self.settings_dialog = SettingsDialog()
        self.settings_dialog.show()
        
        
        

app = QApplication(sys.argv)    
window = MainWindow()
window.showMaximized()
app.exec()
