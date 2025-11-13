
from pathlib import Path
from openpyxl import load_workbook
import re
import random
import os

from EI_protocols_utils.utils.models import ExcelProtocol, WaterMeterProtocol
from EI_protocols_utils.utils.settings import settings
from EI_protocols_utils.utils.constants import *
from utils.exchanges import RequiredFieldsError, RowError
from utils.weather import get_weather, add_weather
from utils.user_info import save_paths, load_paths

class Journal:
    required_fields = [0, 2, 5, 7, 9, 10, 12, 13, 21, 32, 35, 44, 45]
    
    def __init__(self, path: Path):
        self.path = path
        self.workbook = load_workbook(filename=self.path)
        self.wsheet = self.workbook[JOURNAL_WORKSHEET]
        
    def validate_row(self, row: list) -> list:
        # Выводим пропущенные обязательные поля
        missing_fields = []
        for index in self.required_fields:
            if not row[index]:
                missing_fields.append(index)
        
        cells_names = {}
        first_row = self.wsheet[1] # Нумерация в openpyxl с 1
        for index in self.required_fields:
            field_name = first_row[index].value
            cells_names[index] = field_name
        
        if missing_fields:
            # Получаем названия пропущенных полей
            missing_field_names = [cells_names[index] for index in missing_fields]
            # Формируем человекочитаемую строку
            missing_field_list = ", ".join(missing_field_names)
            raise RequiredFieldsError(f"Пропущены поля: {missing_field_list}", missing_fields)

        row[1] = '1' if not row[1] else str(row[1])
        
        # Если не указана погода, то создаем или берем из журнала погоды
        if not row[14] or not row[15] or not row[16]:
            date_str = row[9].strftime("%d.%m.%Y") if type(row[9]) is not str else row[9]
            weather = get_weather(date_str)
            if weather and (not row[14] or not row[15] or not row[16]):
                row[14] = weather["temperature"] + ' °C'
                row[15] = weather["pressure"] + ' кП'
                row[16] = weather["humidity"] + ' %'
                if not get_weather(date_str):
                    add_weather(date_str, row[14], row[15], row[16])  # Обновляем журнал погоды
            else:
                # Если погоды нет, то генерируем случайные значения и сохраняем в журнал погоды
                temperature = str(round(random.uniform(settings.temperatures["min"], settings.temperatures["max"]), 1))
                pressure = str(round(random.uniform(settings.pressure["min"], settings.pressure["max"]), 1))
                humidity = str(round(random.uniform(settings.humdity["min"], settings.humdity["max"]), 1))
                add_weather(date_str, temperature, pressure, humidity)
                row[14] = temperature + ' °C'
                row[15] = pressure + ' кП'
                row[16] = humidity + ' %'

        row[34] = '1' if not row[34] else str(row[34])
        
        
        #
        #   ДОДЕЛАТЬ, БРАТЬ ИЗ ГОТОВОГО ПРОТОКОЛА
        #
        # if not row[41]:
        #     row[41] = f"Поверен в диапазоне расхода (0,03-{round(random.uniform(settings.flow_ranges['min'], settings.flow_ranges['max']), 1)}) м3/ч"

        if not row[47]: row[47] = "Частное лицо"

        return row
        
    def create_protocols(self, from_row: int, to_row: int, to_folder: str) -> list[str]:
        errors = []
        completed = []
        not_completed = []
        for index, row in enumerate(self.wsheet.iter_rows(min_row=from_row, max_row=to_row, values_only=False)):
            try:
                # Преобразуем к значениям для работы
                values = [cell.value for cell in row]

                # for index, value in enumerate(values):
                #     print(f"{index}: {value}")

                # Проверяем и дополняем данные
                self.validate_row(values)
                
                
                
                temperature_str = re.sub(r"[^0-9.]", "", str(values[14]).replace(',', '.'))
                pressure_str = re.sub(r"[^0-9.]", "", str(values[15]).replace(',', '.'))
                humidity_str = re.sub(r"[^0-9.]", "", str(values[16]).replace(',', '.'))
                readings_str = re.sub(r"[^0-9.]", "", str(values[45]).replace(',', '.'))
                
                temperature_float = round(float(temperature_str), 1)
                pressure_float = round(float(pressure_str), 1)
                humidity_float = round(float(humidity_str), 1)
                readings_float = round(float(readings_str), 1)
                
                # Создаем протокол
                protocol = WaterMeterProtocol(
                    dir_path=to_folder,
                    tab_number=settings.tab_numbers.get(values[35], values[0].split('-')[2]),
                    protocol_number=values[0].split('-')[-1],
                    date=values[9].strftime("%d.%m.%Y") if type(values[9]) is not str else values[9],
                    next_date=values[10].strftime("%d.%m.%Y") if type(values[10]) is not str else values[10],
                    SI_numbers=values[12],
                    suitability=False if values[13]=="Непригодно" else True,
                    reasons_for_unsuitability=values[38],
                    name=values[5],
                    number=values[7],
                    register_number=values[2],
                    year=int(values[44]),
                    owner=values[47] if values[47] else "Частное лицо",
                    address=values[32],
                    temperature=temperature_float,
                    pressure=pressure_float,
                    humidity=humidity_float,
                    readings=readings_float,
                    unit_type=values[48]
                )
                
                xlsx_path, pdf_path = protocol.create()
                completed.append((xlsx_path, pdf_path))
                print(f"Создан протокол: {xlsx_path}")
                
                ready_protocol = load_workbook(filename=xlsx_path, data_only=True) # Без data_only=False будут формулы
                ready_wsheet = ready_protocol[WATER_METER_PROTOTOCOL_SEETNAME]
                
                if "Измерения на расходе Qнаиб , л/ч" in str(ready_wsheet["B55"].value):
                    consumption=max(float(ready_wsheet["AC55"].value), float(ready_wsheet["AC54"].value), float(ready_wsheet["AC53"].value))
                elif "Измерения на расходе Qнаиб , л/ч" in ready_wsheet["B59"].value:
                    consumption=max(float(ready_wsheet["AC61"].value), float(ready_wsheet["AC60"].value), float(ready_wsheet["AC59"].value))
                else:
                    raise Exception("Не удалось определить максимальный расход из протокола"\
                        "Убедитесь, что шаблон протокола содержит результаты расхода измерений в B53-B55, или AC59-AC61")
                    
                values[41] = f"Поверен в диапазоне расхода (0,03-{round(consumption, 3)}) м3/ч"
                for i, cell in enumerate(row):
                    cell.value = values[i]
                    
            except Exception as e:
                errors.append(RowError(e, index+from_row))
                not_completed.append(index+from_row)
                raise e
            except RequiredFieldsError as rfe:
                not_completed.append(index+from_row)
                errors.append(RowError(rfe, index+from_row))

        print(f"Выполнены успешно ({len(completed)}): ")
        for item in completed:
            print(f" - {item[0].name}, {item[1].name}")
        print(f"\nС ошибками ({len(errors)}): ")
        for error in errors:
            print(f" - {error}, строка {error.row_number}")
            
        if errors:
            if input("\nСохранить протоколы? (y/n): ").lower() == 'n':
                for item in completed:
                    print(f"Удаляем протоколы: {item[0].name}, {item[1].name}")
                    os.remove(item[0])
                    os.remove(item[1])
                    
        self.workbook.save(filename=self.path)

journal = Journal(path=Path("data/journal.xlsx").resolve())

data = load_paths(filename=settings.user_info_path)
journal_path=data.get('journal_path')
protocols_path=data.get('protocols_path')

if not journal_path: 
    journal_path = input(f"Введите путь до журнала:")
    data['journal_path'] = journal_path
    
else:
    question = input(f"Использовать этот журнал: {journal_path}? Enter если хотите использовать, если нет то введите нужный путь:")
    if question: 
        journal_path=question
        data['journal_path'] = journal_path

if not protocols_path: 
    protocols_path = input(f"Введите путь до папки к готовым протоколам:")
    data['protocols_path'] = protocols_path
else:
    question = input(f"Использовать этот путь: {protocols_path}? Enter если хотите использовать, если нет то введите нужный путь:")
    if question: 
        protocols_path=question
        data['protocols_path'] = protocols_path

save_paths(paths=data, filename=settings.user_info_path)

from_row = int(input(f"С какой строки начать:"))
to_row = int(input(f"На какой закончить:"))

journal.create_protocols(from_row=from_row, to_row=to_row, to_folder=protocols_path)